import os
import sys
import datetime
import openpyxl 

from openpyxl import load_workbook

import mat_ui

import PyQt4 
from PyQt4 import QtGui
from PyQt4 import QtCore
from PyQt4.QtGui import QMessageBox, QInputDialog, QWidget

# *---- FILE PATHS ----* #
PRJ1_PATH = "/path/to/prj1_data_path"
PRJ2_PATH = "/path/to/prj2_data_path"
# *--------------------* #

wb = load_workbook('./template.xlsx')       # first load a workbook
lTotalSheets = wb.get_sheet_names()

SuffixCount = 1
rowCountIndex = 9

item = None
description = None

class MocapAssistantTool(QtGui.QMainWindow, mat_ui.Ui_MainWindow):
    def __init__(self, parent=None):
        super(MocapAssistantTool, self).__init__(parent)
        self.showUI()

    def showUI(self):
        self.setupUi(self)
        self.btn_about_this_tool.clicked.connect(self.callback_AboutWindow)
        self.btn_set_actor_choice.clicked.connect(self.callback_SetActorField)
        self.btn_set_project_choice.clicked.connect(self.callback_SetProjectField)
        self.btn_saveas.clicked.connect(self.callback_SaveProjectAs)
        self.btn_savefile.clicked.connect(self.callback_SaveProject)
        self.btn_createtake.clicked.connect(self.callback_MocapRowEntry)
        self.btn_removetake.clicked.connect(self.callback_RemoveMocapRowEntry)

    def callback_SetActorField(self):
        if self.combobox_actor_choice.currentText() == '-- Select Actor --':
            self.le_outputfield.setText('No Actor Selected!')
        else:
            self.ModifyFieldValue('Actor', str(self.combobox_actor_choice.currentText()))
            self.le_outputfield.setText('Actor set to ' + self.combobox_actor_choice.currentText() + '!')
            SaveFile('template.xlsx')
            print('Set project to '+ str(self.combobox_actor_choice.currentText()) +' & saved!')

    def callback_SetProjectField(self):
        if self.combobox_project_choice.currentText() == '-- Select Project --':
            self.le_outputfield.setText('None Selected!')
        else:
            self.ModifyFieldValue('Directory', str(self.combobox_project_choice.currentText()))
            self.le_outputfield.setText("Path Set: " + self.combobox_project_choice.currentText())
            self.callback_UpdateDirectoryPath()
            SaveFile('template.xlsx')
            print('Set to '+ str(self.combobox_project_choice.currentText()) +' & saved!')

    def callback_SaveProject(self):
        self.callback_PrintTimeStamp()
        SaveFile('./template.xlsx')
        self.le_outputfield.setText('Saved!')

    def callback_SaveProjectAs(self):
        CheckCurrentDirectory()
        self.callback_PrintTimeStamp()
        SaveFile(str(self.le_filenamefield.text())+'.xlsx')
        print(str(self.le_filenamefield.text()))
        self.le_outputfield.setText('Saved as: '+ str(self.le_filenamefield.text()))

    def callback_UpdateDirectoryPath(self):
        if sheet['A4'].value == 'Project 1':
            SetActiveSheet(0)
            self.ModifyFieldValue('ProjectPath', str(PRJ1_PATH))
        if sheet['A4'].value == 'Project 2':
            SetActiveSheet(1)
            self.ModifyFieldValue('ProjectPath', str(PRJ2_PATH))

    def testFunction(self):                   # test function for debugging, remove later.
        print('Button is working!')

    def ModifyFieldValue(self, pWhatField, pWhatData):
        sheet = wb.get_sheet_by_name('Project 1')
        if pWhatField == 'Date':
            DATEFIELD.value = pWhatData
        if pWhatField == 'Actor':
            ACTORFIELD.value = pWhatData
        if pWhatField == 'Directory':
            DIRFIELD.value = pWhatData
        if pWhatField == 'ProjectPath':
            PRJFIELD.value = pWhatData
        SaveFile('template.xlsx')

    def callback_MocapRowEntry(self):
        # https://stackoverflow.com/questions/7907928/openpyxl-check-for-empty-
        global item
        global description
        global SuffixCount
        global rowCountIndex

        sheet = wb.get_sheet_by_name('Project 1')

        sheet['A'+str(rowCountIndex)].value = str(SuffixCount)
        sheet['B'+str(rowCountIndex)].value = 'MVN_Take_'+ str(SuffixCount)
        sheet['C'+str(rowCountIndex)].value = 'FBX_'+str(SuffixCount)+'.fbx'

        # for selecting which level the mocap was captured on #
        items = ('No-Level', 'Single-Level', 'Multi-Level')
        item, ok = QInputDialog.getItem(self, "Select Which Level", "Capture-Type", items, 0, False)
        if ok and item:
            sheet['D'+str(rowCountIndex)].value = str(item)

        # for entering the description for the mocap take #
        description, ok = QInputDialog.getText(self, "Take Description", "Enter a Description for this Take:")
        if ok and description:
            sheet['E'+str(rowCountIndex)].value = str(description)
            self.le_outputfield.setText('Entry '+ str(SuffixCount)+' Entered!')
        print('|| Entry: ' + str(SuffixCount) + ' | Level-Type: ' + str(item) + ' | Description: ' + str(description) + ' ||')

        SaveFile('template.xlsx')
        SuffixCount += 1
        rowCountIndex += 1

    def callback_RemoveMocapRowEntry(self):
        global item
        global description
        global SuffixCount
        global rowCountIndex

        sheet= wb.get_sheet_by_name('Project 1')
        if rowCountIndex == 9:
            self.le_outputfield.setText('Cant delete!')
            print('Nothing to delete!')
        else:
            rowCountIndex -= 1
            SuffixCount -= 1
            sheet['A'+str(rowCountIndex)].value = " "
            sheet['B'+str(rowCountIndex)].value = " "
            sheet['C'+str(rowCountIndex)].value = " "
            sheet['D'+str(rowCountIndex)].value = " "
            sheet['E'+str(rowCountIndex)].value = " "
            self.le_outputfield.setText('Entry '+ str(SuffixCount)+' Deleted!')
            print('Removed ---- || Entry: ' + str(SuffixCount) + ' ||')
        SaveFile('template.xlsx')

    def callback_AboutWindow(self):
        aboutwindow = QMessageBox()
        aboutwindow.setText('Built with Python, openpyxl, PyQt4')
        aboutwindow.setInformativeText("<a href='http://google.com'>goto/mandl/</a>")
        aboutwindow.setWindowTitle('About')
        aboutwindow.setStandardButtons(QMessageBox.Ok)
        callwindow = aboutwindow.exec_()

    def callback_PrintTimeStamp(self):
        today = datetime.datetime.now()
        self.ModifyFieldValue('Date', str(today.date()))
        print(today.date())


# *---- DATA FRAME ----* #
sheet = wb.get_sheet_by_name('Project 1')
DATEFIELD = sheet['E1']
ACTORFIELD = sheet['E2']
DIRFIELD = sheet['A4']
PRJFIELD = sheet['B4']
# *--------------------* #

def CheckCurrentDirectory():
    cwd = os.getcwd()                       # check current working directory (cwd)
    print(cwd)
    cwdfiles = os.listdir('.')              # prints a list of whats in the current working directory
    print(cwdfiles)
#
# def CreateNewSheets(pSheetName):
#     lTotalSheets = wb.get_sheet_names()
#     print(len(lTotalSheets))
#     newSheet = wb.create_sheet(pSheetName)
#     lTotalSheets.append(newSheet)
#     SaveFile()
#     return lTotalSheets

def CloseFile():
    try:
        os.system('TASKKILL /IM Excel.exe')
    except Exception:
        print(str(e))

def SaveFile(pFileNameToSave):
    wb.save(filename = pFileNameToSave)

def SetActiveSheet(pIndex):
    allsheets = wb.sheetnames
    activesheet = wb[allsheets[pIndex]]
    SuffixCount = 1
    rowCountIndex = 9
    print(activesheet)

def main():
    app = QtGui.QApplication(sys.argv)
    form = MocapAssistantTool()
    form.show()
    app.exec_()

if __name__ == '__main__':
    main()
